import calendar
import locale
import os

from flask import app
from openpyxl import load_workbook, workbook
from openpyxl.workbook import Workbook
from openpyxl.styles import Border, Side, Font, Alignment



class Tenant:
    def __init__(self, t_dictionary):
        for key in t_dictionary:
            setattr(self, key, t_dictionary[key])



def fetch_input_workbook(filename):
    """
    Imports the Excel workbook to be converted
    :return:
    """
    try:
        UPLOAD_FOLDER = './files/uploads'
        input_file = os.path.join(UPLOAD_FOLDER, filename)
        wb = load_workbook(filename=input_file)
        return wb
    except FileNotFoundError as e :
        print("Input file not found", e)
        exit(1)


def set_up_output_workbook(name, month, year):
    """
     Open the output Excel worksheet (month) in the workbook (name + "Rent Roll " + year)
     If the worksheet isn't found (and it shouldn't be), create the worksheet.
     Also, remove the default "Sheet" tab in the created workbook.
     :return:
     """
    try:
        output_file = "%s Rent Roll %s.xlsx" % (name, year)
        output_wb = Workbook(write_only=False)
        output_wb.title = output_file
        output_sheets = output_wb.sheetnames

        if month not in output_sheets:
            output_wb.create_sheet(month)
        output_wb.save(output_file)

        output_sheets = output_wb.sheetnames

        if 'Sheet' in output_sheets:
            del output_wb['Sheet']

        output_wb.active = output_wb[month]
        active_sheet = output_wb.active

        #PrintPageSetup(output_wb, orientation="landscape", paperSize=1, usePrinterDefaults="FALSE")


        output_wb.save(output_file)
        return output_wb
    except Exception as e:
        print(e)
        print('Trying to write to output_file ' + str(output_file) + ' and something went wrong.')
        exit(2)


def write_top_info(output_wkbk, month, property_dict, row):
    """
    Adds top of page information
    :param output_wkbk:  Output workbook
    :param month: The month taken from the input file
    :param property_dict: Dictionary of property identification variations
    :param row: The row to start on. Should always be 0.
    :return:
    """
    output_wkbk.active = output_wkbk[month]
    active_sheet = output_wkbk.active

    # write the headings
    active_sheet.merge_cells(('A' + str(row) + ":" + 'B' + str(row)))
    active_sheet['A' + str(row)] = property_dict['page_title']
    row = row + 1

    active_sheet.merge_cells(('A' + str(row) + ":" + 'B' + str(row)))
    active_sheet['A' + str(row)] = property_dict['property_name_short'] + '-' + property_dict['property_name_UC']
    row = row + 1

    active_sheet['A' + str(row)] = "AS OF"
    active_sheet['B' + str(row)] = property_dict['month']


def write_column_headings(output_workbook, text_month, row, property_dict):
    """
    Write the column headings to simulate page breaks
    :param pagebreak:
    :param output_workbook:
    :param text_month:
    :param row:
    :param property_dict:
    :return:
    """
    output_workbook.active = output_workbook[text_month]
    active_sheet = output_workbook.active
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    headings = property_dict['list_column_headers']


    # create the border around the headings and set dimensions
    for cell in range(1, 12):
        active_sheet.cell(row=row, column=cell).border = thin_border
        active_sheet.cell(row=row, column=cell).alignment = Alignment(wrap_text=True)
        active_sheet.column_dimensions['B'].width = 18
        active_sheet.column_dimensions['C'].width = 7
        active_sheet.column_dimensions['D'].width = 7
        active_sheet.column_dimensions['E'].width = 7
        active_sheet.column_dimensions['K'].width = 20
    active_sheet.row_dimensions[5].height = 35

    active_sheet['A' + str(row)] = headings[0]
    active_sheet['B' + str(row)] = headings[1]
    active_sheet['C' + str(row)] = headings[2]
    active_sheet['D' + str(row)] = headings[3]
    active_sheet['E' + str(row)] = headings[4]
    active_sheet['F' + str(row)] = headings[5]
    active_sheet['G' + str(row)] = headings[6]
    active_sheet['H' + str(row)] = headings[7]
    active_sheet['I' + str(row)] = headings[8]
    active_sheet['J' + str(row)] = headings[9]
    active_sheet['K' + str(row)] = headings[10]


def write_data(output_wkbk, month, tenants, property_info):
    """
    Write all the data in the tenants list, breaking it up so that the
    column heading repeat at intervals (simulating page breaks)
    :param output_wkbk:
    :param month:
    :param tenants:
    :param property_info:
    :return:
    """
    active_sheet = output_wkbk[month]
    row = 8
    for tenant in tenants:
        active_sheet['A' + str(row)] = tenant.tenant_code
        active_sheet['B' + str(row)] = tenant.tenant_name

        if hasattr(tenant, 'rent'):
            active_sheet['C' + str(row)] = tenant.rent
            if hasattr(tenant, 'mgrdisc'):
                active_sheet['C' + str(row)] = tenant.rent + tenant.mgrdisc

        if hasattr(tenant, 'addocc'):
            active_sheet['D' + str(row)] = tenant.addocc

        if hasattr(tenant, 'parking'):
            active_sheet['E' + str(row)] = tenant.parking
            if hasattr(tenant, 'mngrpkds'):
                active_sheet['E' + str(row)] = tenant.parking + tenant.mngrpkds

        if hasattr(tenant, 'storage'):
            active_sheet['F' + str(row)] = tenant.storage

        active_sheet['G' + str(row)] = '=SUM(C' + str(row) + ' + D' + str(row) + ' + E' + str(row) + ' + F' + str(
            row) + ')'
        row = row + 1

        if row % 32 == 0:
            write_column_headings(output_wkbk, month, row, property_info)
            row = row + 2
    return row


def generate_totals_section(output_workbook, text_month, output_row):
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    medium_border = Border(top=Side(style='medium'))

    active_sheet = output_workbook[text_month]
    active_sheet['B' + str(output_row)] = "Total"
    active_sheet['B' + str(output_row)].font = Font(color="000000", bold=True)

    # active_sheet['C' + str(output_row)] = tenant_total
    active_sheet['C' + str(output_row)] = '=SUM(C1:C' + str(output_row - 1) + ')'
    active_sheet['C' + str(output_row)].font = Font(color="000000", bold=True)

    # active_sheet['D' + str(output_row)] = total_occ
    active_sheet['D' + str(output_row)] = '=SUM(D1:D' + str(output_row - 1) + ')'
    active_sheet['D' + str(output_row)].font = Font(color="000000", bold=True)

    # active_sheet['E' + str(output_row)] = total_parking
    active_sheet['E' + str(output_row)] = '=SUM(E1:E' + str(output_row - 1) + ')'
    active_sheet['E' + str(output_row)].font = Font(color="000000", bold=True)

    # active_sheet['F' + str(output_row)] = total_storage
    active_sheet['F' + str(output_row)] = '=SUM(F1:F' + str(output_row - 1) + ')'
    active_sheet['F' + str(output_row)].font = Font(color="000000", bold=True)

    # active_sheet['G' + str(output_row)] = total_rent
    active_sheet['G' + str(output_row)] = '=SUM(G1:G' + str(output_row - 1) + ')'
    active_sheet['G' + str(output_row)].font = Font(color="000000", bold=True)

    for cell in range(2, 12):
        active_sheet.cell(row=output_row, column=cell).border = medium_border

    output_row = output_row + 2
    active_sheet['H' + str(output_row)] = "Please enter totals for above 3 columns"
    active_sheet['H' + str(output_row)].font = Font(color="000000", italic=True)
    output_row = output_row + 4
    total_labels = ['Less Under Paid',
                    'Add Over Paid',
                    'Total Receipts',
                    'Total Expected',
                    'Total Received',
                    'Total Underpaid',
                    'Total Overpaid']
    for label in total_labels:
        active_sheet['C' + str(output_row)] = label

        for cell in range(8, 12):
            active_sheet.cell(row=output_row, column=cell).border = thin_border
        if label == 'Total Receipts':
            active_sheet['H' + str(output_row)] = '=G' + str(output_row - 8)
            active_sheet['H' + str(output_row)].font = Font(color="000000", bold=True)
            output_row = output_row + 1

        output_row = output_row + 1


def process_sheets(wkb, wk_sheets):
    """
    Grabs the latest worksheet in a workbook to process the data
    After talking with Jaime on Dec 17, it was clarified that
    there will only be one sheet in the notebook.

    I'm leaving the code as is in case, at a later date, they
    decide to add more sheets.  It still works for workbooks with just one sheet.
    :param wkb:
    :param wk_sheets:
    :return:
    """
    number_sheets = len(wk_sheets)
    latest_sheet = wk_sheets[number_sheets - 1]
    active_worksheet = wkb[latest_sheet]

    total_rent = 0
    total_occ = 0
    total_parking = 0
    total_storage = 0
    total_total = 0

    """
        property_designation -> Type = Name
        e.g.
        property_designation -> Property = Carmen Manor
        type -> Property
        property_name -> Carmen Manor
        property_name_UC -> CARMEN MANOR
        property_name_shortened -> Carmen
        """
    property_designation = active_worksheet['A2'].value
    property_name = property_designation[property_designation.index("=") + 3:]
    property_name_UC = property_name.upper()
    property_name_shortened = property_name[0:property_name.index(" ")]
    type = property_designation[0:property_designation.index("=") - 1]

    """
        month_year = 'Month = 01/2023'
        month = '01'
        month = 'January'
        year = 2023				
    """
    locale.setlocale(locale.LC_ALL, 'en_CA')
    as_of_date = active_worksheet['A3'].value  # e.g. As Of = 01/01/2023
    day_of_month = as_of_date[as_of_date.index("/") + 1:13]  # e.g. 01
    month_year = active_worksheet['A4'].value  # e.g. Month = 01/2023
    month = int(month_year[month_year.index("=") + 2:month_year.index('/')])
    text_month = calendar.month_name[month]
    year = month_year[month_year.index("/") + 1:]

    output_workbook = set_up_output_workbook(property_name, text_month, year)

    # property_info holds all the page headings and column headings

    property_info = {}
    property_info['property_name_short'] = property_name_shortened
    property_info['property_name_UC'] = property_name_UC
    property_info['month'] = text_month
    property_info['page_title'] = 'Rent Collection Form'
    property_info['property_designation'] = property_name_UC + ' - ' + property_name.upper()
    property_info['as_of_date'] = as_of_date
    property_info['list_column_headers'] = ["Tenant Code", "Tenant Name", "Rent",
                                            "Add Occ", "Parking", "Storage Locker",
                                            "Total Charges", "Amount Received",
                                            "Under Paid", "Over Paid", "Manager's Comments"]

    row = 1
    write_top_info(output_workbook, text_month, property_info, row)

    row = 5
    write_column_headings(output_workbook, text_month, row, property_info)

    output_workbook.save(output_workbook.title)
    desired_value = None
    row = 8  # will hold the current input row - starts at row 8
    count = 0
    row_count = 1  # Will hold the number of rows containing data before the total row
    output_row = 7
    cost_category = ''
    categories = {}
    total = 0
    new_tenant_flag = False
    tenants = []
    tenant_dict = {}
    future_tenants = []

    # Find the summary section so we know when to stop trying to read in tenants
    while active_worksheet['A' + str(row_count)].value != 'Total':
        row_count = row_count + 1
    total_row = row_count

    # Now, process the tenants.  For each tenant, pull all the costs and create a Tenant object
    # and append to the list of tenants
    while row < total_row:
        tenant_code_cell = active_worksheet['A' + str(row)].value

        if tenant_code_cell is not None and tenant_code_cell != 'Future Tenants/Applicants':

            tenant_code = tenant_code_cell

            tenant_name = active_worksheet['C' + str(row)].value
            if active_worksheet['E' + str(row)].value != None:
                cost_category = active_worksheet['E' + str(row)].value

            while cost_category != 'Total':
                """
                There are multiple mgrdisc fields and multiple parking fields in some of the files. 
                This is a generic catch for all multiple fields. 
                """
                if cost_category in categories:
                    categories[cost_category] = categories[cost_category] + (active_worksheet['F' + str(row)]).value
                else:
                    categories[cost_category] = (active_worksheet['F' + str(row)]).value
                row = row + 1
                cost_category = active_worksheet['E' + str(row)].value

            categories['Total'] = sum(categories.values())

            # Build out dictionary of tenant info
            tenant_dict['tenant_code'] = tenant_code
            tenant_dict['tenant_name'] = tenant_name

            for k, v in categories.items():
                tenant_dict[k] = v


            # Create a tenant instance and add to list.
            # We are doing it this way so we can sort the list by tenant_code
            # and future/new tenants will be inserted into the list at the correct location.

            tenant = Tenant(tenant_dict)

            tenants.append(tenant)

            if new_tenant_flag:
                future_tenants.append(tenant)

            # Update running totals
            if 'rent' in categories.keys():
                total_rent = total_rent + categories['rent']

            if 'addocc' in categories.keys():
                total_occ = total_occ + categories['addocc']

            if 'parking' in categories.keys():
                total_parking = total_parking + categories['parking']

            if 'mgr' in categories.keys():
                total_rent = total_rent - categories['mgr']

            if 'storage' in categories.keys():
                total_storage = total_storage + categories['storage']

            categories.clear()
            tenant_dict.clear()
            cost_category = ''

        if tenant_code_cell == 'Future Tenants/Applicants':
            new_tenant_flag = True

        row = row + 1

    tenants.sort(key=lambda x: x.tenant_code)
    # Remove the duplicates
    # e.g.
    # PP0716	VACANT					                0
    # PP0716	Sue Smith	        1300				1300
    # Sue is a future/new tenant.  Replace the VACANT row with the other row.

    for new_tenant in future_tenants:
        for est_tenant in tenants:
            if est_tenant.tenant_code == new_tenant.tenant_code:
                tenants.remove(est_tenant)
                # print("Replacing ", est_tenant.__dict__, "with ", new_tenant.__dict__)

    last_row = write_data(output_workbook, text_month, tenants, property_info)

    # Now write the totals
    generate_totals_section(output_workbook, text_month, last_row)
    DOWNLOAD_FOLDER = './files/downloads'
    output_workbook.save(os.path.join(DOWNLOAD_FOLDER, output_workbook.title))
    return output_workbook.title




